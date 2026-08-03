# backend/src/presentation/http/routers/reports.py
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import JSONResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, case
from datetime import datetime
from typing import Optional, Any
import os
import json
import math

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from src.infrastructure.db.init_db import get_db

REPORTS_DIR = "uploads/reports"
os.makedirs(REPORTS_DIR, exist_ok=True)

router = APIRouter(prefix="/reports", tags=["reports"])


def safe_decimal_to_float(value: Optional[Any], default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        if isinstance(value, (int, float)):
            return float(value)
        return default
    except (TypeError, ValueError):
        return default


def safe_str(value: Optional[Any], default: str = "") -> str:
    if value is None:
        return default
    try:
        return str(value)
    except (TypeError, ValueError):
        return default


def safe_isoformat(value: Optional[Any]) -> Optional[str]:
    if value is None:
        return None
    try:
        if hasattr(value, 'isoformat') and callable(getattr(value, 'isoformat')):
            return value.isoformat()
        return str(value)
    except (AttributeError, ValueError):
        return None


def get_worksheet(wb: Workbook):
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Sheet1")
    return ws


def _style_header(ws, row, col, value):
    """Стилизация заголовка"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


def _style_cell(ws, row, col, value, bold=False):
    """Стилизация ячейки"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold)
    thin = Side(border_style="thin", color="999999")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


# ======================================================================
# ОТЧЁТ ПО АКТИВАМ
# ======================================================================

@router.get("/asset-report")
async def get_asset_report(db: Session = Depends(get_db)):
    """Отчет по активам (JSON) — расширенный"""
    try:
        from src.infrastructure.db.models.asset import Asset
        from src.infrastructure.db.models.asset_type_config import AssetTypeConfig

        assets = db.query(Asset).filter(Asset.is_active == True).all()
        type_configs = {tc.code: tc for tc in db.query(AssetTypeConfig).all()}

        # Общая статистика
        total_count = len(assets)
        total_purchase = sum(safe_decimal_to_float(a.purchase_price) for a in assets)
        total_current = sum(safe_decimal_to_float(a.current_value) for a in assets)
        total_depreciation = total_purchase - total_current

        # По типам
        type_breakdown = {}
        for a in assets:
            t = a.asset_type or "unknown"
            if t not in type_breakdown:
                tc = type_configs.get(t)
                type_breakdown[t] = {
                    "code": t,
                    "name": tc.name if tc else t,
                    "icon": tc.icon if tc else "📦",
                    "count": 0,
                    "total_purchase": 0.0,
                    "total_current": 0.0,
                }
            type_breakdown[t]["count"] += 1
            type_breakdown[t]["total_purchase"] += safe_decimal_to_float(a.purchase_price)
            type_breakdown[t]["total_current"] += safe_decimal_to_float(a.current_value)

        # По статусам
        status_map = {"active": "Активен", "maintenance": "На ремонте", "reserved": "В резерве",
                      "decommissioned": "Выведен", "lost": "Утерян", "written_off": "Списан"}
        status_breakdown = {}
        for a in assets:
            s = a.status or "unknown"
            if s not in status_breakdown:
                status_breakdown[s] = {"status": s, "label": status_map.get(s, s), "count": 0, "total_current": 0.0}
            status_breakdown[s]["count"] += 1
            status_breakdown[s]["total_current"] += safe_decimal_to_float(a.current_value)

        # По подразделениям
        dept_breakdown = {}
        for a in assets:
            d = a.department_code or "Без подразделения"
            if d not in dept_breakdown:
                dept_breakdown[d] = {"department": d, "count": 0, "total_purchase": 0.0, "total_current": 0.0}
            dept_breakdown[d]["count"] += 1
            dept_breakdown[d]["total_purchase"] += safe_decimal_to_float(a.purchase_price)
            dept_breakdown[d]["total_current"] += safe_decimal_to_float(a.current_value)

        # Активы, требующие внимания (на ремонте, устаревшие)
        needs_repair = [a for a in assets if a.status == "maintenance"]
        low_value = [a for a in assets if safe_decimal_to_float(a.current_value) < safe_decimal_to_float(a.purchase_price) * 0.1 and safe_decimal_to_float(a.purchase_price) > 0]

        assets_list = []
        for asset in assets:
            assets_list.append({
                "id": getattr(asset, 'id', None),
                "inventory_number": safe_str(getattr(asset, 'inventory_number', None)),
                "name": safe_str(getattr(asset, 'name', None)),
                "model": safe_str(getattr(asset, 'model', None)),
                "asset_type": safe_str(getattr(asset, 'asset_type', None)),
                "status": safe_str(getattr(asset, 'status', None)),
                "current_value": safe_decimal_to_float(getattr(asset, 'current_value', None)),
                "purchase_price": safe_decimal_to_float(getattr(asset, 'purchase_price', None)),
                "department_code": safe_str(getattr(asset, 'department_code', None)),
                "responsible_person": safe_str(getattr(asset, 'responsible_person', None)),
                "location_address": safe_str(getattr(asset, 'location_address', None)),
                "created_at": safe_isoformat(getattr(asset, 'created_at', None)),
            })

        return JSONResponse(content={
            "title": "Отчет по активам",
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "total_count": total_count,
                "total_purchase_value": round(total_purchase, 2),
                "total_current_value": round(total_current, 2),
                "total_depreciation": round(total_depreciation, 2),
                "depreciation_percent": round((total_depreciation / total_purchase * 100) if total_purchase > 0 else 0, 1),
                "needs_repair_count": len(needs_repair),
                "low_value_count": len(low_value),
            },
            "type_breakdown": sorted(type_breakdown.values(), key=lambda x: x["count"], reverse=True),
            "status_breakdown": sorted(status_breakdown.values(), key=lambda x: x["count"], reverse=True),
            "department_breakdown": sorted(dept_breakdown.values(), key=lambda x: x["count"], reverse=True),
            "assets": assets_list,
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@router.get("/asset-report/export")
async def export_asset_report(
    format: str = Query("excel", description="Формат: excel, pdf, json"),
    db: Session = Depends(get_db),
):
    """Экспорт отчета по активам"""
    try:
        report_data = await get_asset_report(db)
        data = json.loads(report_data.body.decode())

        if format == "excel":
            wb = Workbook()
            ws = get_worksheet(wb)
            ws.title = "Сводка"

            # Заголовок
            ws.merge_cells('A1:F1')
            ws['A1'] = data['title']
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal="center")

            row = 3
            s = data['summary']
            for label, val in [("Всего активов", s['total_count']),
                               ("Общая стоимость покупки", f"{s['total_purchase_value']:,.2f} ₽"),
                               ("Текущая стоимость", f"{s['total_current_value']:,.2f} ₽"),
                               ("Амортизация", f"{s['total_depreciation']:,.2f} ₽ ({s['depreciation_percent']}%)"),
                               ("Требуют ремонта", s['needs_repair_count']),
                               ("Сильно устарели", s['low_value_count'])]:
                _style_header(ws, row, 1, label)
                _style_cell(ws, row, 2, val)
                row += 1

            # По типам
            row += 2
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1, value="По типам активов").font = Font(bold=True, size=14)
            row += 1
            for col, h in enumerate(["Тип", "Количество", "Стоимость покупки", "Текущая стоимость", "Доля"], 1):
                _style_header(ws, row, col, h)
            row += 1
            for t in data['type_breakdown']:
                _style_cell(ws, row, 1, f"{t['icon']} {t['name']}")
                _style_cell(ws, row, 2, t['count'])
                _style_cell(ws, row, 3, f"{t['total_purchase']:,.2f} ₽")
                _style_cell(ws, row, 4, f"{t['total_current']:,.2f} ₽")
                _style_cell(ws, row, 5, f"{t['count'] / s['total_count'] * 100:.1f}%" if s['total_count'] > 0 else "0%")
                row += 1

            # По статусам
            row += 2
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row=row, column=1, value="По статусам").font = Font(bold=True, size=14)
            row += 1
            for col, h in enumerate(["Статус", "Количество", "Стоимость", "Доля"], 1):
                _style_header(ws, row, col, h)
            row += 1
            for s_item in data['status_breakdown']:
                _style_cell(ws, row, 1, s_item['label'])
                _style_cell(ws, row, 2, s_item['count'])
                _style_cell(ws, row, 3, f"{s_item['total_current']:,.2f} ₽")
                _style_cell(ws, row, 4, f"{s_item['count'] / s['total_count'] * 100:.1f}%" if s['total_count'] > 0 else "0%")
                row += 1

            # По подразделениям
            row += 2
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1, value="По подразделениям").font = Font(bold=True, size=14)
            row += 1
            for col, h in enumerate(["Подразделение", "Количество", "Стоимость покупки", "Текущая стоимость", "Доля"], 1):
                _style_header(ws, row, col, h)
            row += 1
            for d in data['department_breakdown']:
                _style_cell(ws, row, 1, d['department'])
                _style_cell(ws, row, 2, d['count'])
                _style_cell(ws, row, 3, f"{d['total_purchase']:,.2f} ₽")
                _style_cell(ws, row, 4, f"{d['total_current']:,.2f} ₽")
                _style_cell(ws, row, 5, f"{d['count'] / s['total_count'] * 100:.1f}%" if s['total_count'] > 0 else "0%")
                row += 1

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 15

            filename = f"asset_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(REPORTS_DIR, filename)
            wb.save(filepath)
        else:
            filename = f"asset_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            filepath = os.path.join(REPORTS_DIR, filename)
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

        return FileResponse(path=filepath, filename=filename, media_type="application/octet-stream")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


# ======================================================================
# ОТЧЁТ ПО АМОРТИЗАЦИИ
# ======================================================================

@router.get("/depreciation-report")
async def get_depreciation_report(db: Session = Depends(get_db)):
    """Отчет по амортизации (JSON) — расширенный"""
    try:
        from src.infrastructure.db.models.asset import Asset
        from src.infrastructure.db.models.asset_type_config import AssetTypeConfig

        type_configs = {tc.code: tc for tc in db.query(AssetTypeConfig).all()}
        assets = db.query(Asset).filter(Asset.is_active == True).all()

        total_purchase = sum(safe_decimal_to_float(a.purchase_price) for a in assets)
        total_current = sum(safe_decimal_to_float(a.current_value) for a in assets)
        total_depreciation = total_purchase - total_current

        # Распределение по износу
        wear_levels = {"high": {"label": "Высокий (>80%)", "count": 0, "value": 0.0},
                       "medium": {"label": "Средний (50-80%)", "count": 0, "value": 0.0},
                       "low": {"label": "Низкий (<50%)", "count": 0, "value": 0.0},
                       "none": {"label": "Нет данных", "count": 0, "value": 0.0}}

        for a in assets:
            purchase = safe_decimal_to_float(a.purchase_price)
            current = safe_decimal_to_float(a.current_value)
            if purchase > 0:
                wear = (purchase - current) / purchase * 100
                if wear >= 80:
                    wear_levels["high"]["count"] += 1
                    wear_levels["high"]["value"] += current
                elif wear >= 50:
                    wear_levels["medium"]["count"] += 1
                    wear_levels["medium"]["value"] += current
                else:
                    wear_levels["low"]["count"] += 1
                    wear_levels["low"]["value"] += current
            else:
                wear_levels["none"]["count"] += 1
                wear_levels["none"]["value"] += current

        # Активы, у которых износ > 80% (пора списывать)
        high_wear_assets = []
        for a in assets:
            purchase = safe_decimal_to_float(a.purchase_price)
            current = safe_decimal_to_float(a.current_value)
            if purchase > 0 and (purchase - current) / purchase * 100 >= 80:
                tc = type_configs.get(a.asset_type or "")
                high_wear_assets.append({
                    "id": a.id,
                    "inventory_number": safe_str(a.inventory_number),
                    "name": safe_str(a.name),
                    "type": tc.name if tc else a.asset_type,
                    "purchase_price": purchase,
                    "current_value": current,
                    "wear_percent": round((purchase - current) / purchase * 100, 1),
                    "department": safe_str(a.department_code),
                })

        # По типам с износом
        type_wear = {}
        for a in assets:
            t = a.asset_type or "unknown"
            if t not in type_wear:
                tc = type_configs.get(t)
                type_wear[t] = {"code": t, "name": tc.name if tc else t, "icon": tc.icon if tc else "📦",
                                "count": 0, "purchase": 0.0, "current": 0.0}
            type_wear[t]["count"] += 1
            type_wear[t]["purchase"] += safe_decimal_to_float(a.purchase_price)
            type_wear[t]["current"] += safe_decimal_to_float(a.current_value)

        for tw in type_wear.values():
            tw["depreciation_amount"] = round(tw["purchase"] - tw["current"], 2)
            tw["depreciation_percent"] = round(tw["depreciation_amount"] / tw["purchase"] * 100, 1) if tw["purchase"] > 0 else 0

        return JSONResponse(content={
            "title": "Отчет по амортизации",
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "total_assets": len(assets),
                "total_purchase_value": round(total_purchase, 2),
                "total_current_value": round(total_current, 2),
                "total_depreciation": round(total_depreciation, 2),
                "avg_depreciation_percent": round(total_depreciation / total_purchase * 100, 1) if total_purchase > 0 else 0,
            },
            "wear_levels": list(wear_levels.values()),
            "high_wear_assets": sorted(high_wear_assets, key=lambda x: x["wear_percent"], reverse=True),
            "type_wear": sorted(type_wear.values(), key=lambda x: x["depreciation_amount"], reverse=True),
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@router.get("/depreciation-report/export")
async def export_depreciation_report(
    format: str = Query("excel", description="Формат: excel, pdf, json"),
    db: Session = Depends(get_db),
):
    """Экспорт отчета по амортизации"""
    try:
        report_data = await get_depreciation_report(db)
        data = json.loads(report_data.body.decode())

        if format == "excel":
            wb = Workbook()
            ws = get_worksheet(wb)
            ws.title = "Амортизация"

            ws.merge_cells('A1:E1')
            ws['A1'] = data['title']
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal="center")

            row = 3
            s = data['summary']
            for label, val in [("Всего активов", s['total_assets']),
                               ("Первоначальная стоимость", f"{s['total_purchase_value']:,.2f} ₽"),
                               ("Текущая стоимость", f"{s['total_current_value']:,.2f} ₽"),
                               ("Общая амортизация", f"{s['total_depreciation']:,.2f} ₽"),
                               ("Средний износ", f"{s['avg_depreciation_percent']}%")]:
                _style_header(ws, row, 1, label)
                _style_cell(ws, row, 2, val)
                row += 1

            row += 2
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row=row, column=1, value="Уровень износа").font = Font(bold=True, size=14)
            row += 1
            for col, h in enumerate(["Уровень", "Количество", "Остаточная стоимость", "Доля"], 1):
                _style_header(ws, row, col, h)
            row += 1
            for wl in data['wear_levels']:
                _style_cell(ws, row, 1, wl['label'])
                _style_cell(ws, row, 2, wl['count'])
                _style_cell(ws, row, 3, f"{wl['value']:,.2f} ₽")
                _style_cell(ws, row, 4, f"{wl['count'] / s['total_assets'] * 100:.1f}%" if s['total_assets'] > 0 else "0%")
                row += 1

            row += 2
            ws.merge_cells(f'A{row}:F{row}')
            ws.cell(row=row, column=1, value="Активы с высоким износом").font = Font(bold=True, size=14)
            row += 1
            for col, h in enumerate(["Инв. номер", "Название", "Тип", "Цена", "Остаток", "Износ"], 1):
                _style_header(ws, row, col, h)
            row += 1
            for ha in data['high_wear_assets']:
                for col, val in enumerate([ha['inventory_number'], ha['name'], ha['type'],
                                           f"{ha['purchase_price']:,.2f} ₽", f"{ha['current_value']:,.2f} ₽",
                                           f"{ha['wear_percent']}%"], 1):
                    _style_cell(ws, row, col, val)
                row += 1

            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col_letter].width = 22

            filename = f"depreciation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(REPORTS_DIR, filename)
            wb.save(filepath)
        else:
            filename = f"depreciation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            filepath = os.path.join(REPORTS_DIR, filename)
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

        return FileResponse(path=filepath, filename=filename, media_type="application/octet-stream")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


# ======================================================================
# ИНВЕНТАРИЗАЦИОННЫЙ ОТЧЁТ
# ======================================================================

@router.get("/inventory-report")
async def get_inventory_report(db: Session = Depends(get_db)):
    """Инвентаризационный отчет — расширенная статистика"""
    try:
        from src.infrastructure.db.models.asset import Asset
        from src.infrastructure.db.models.asset_type_config import AssetTypeConfig
        from src.infrastructure.db.models.department import Department, Room
        from src.infrastructure.db.models.employee import Employee

        type_configs = {tc.code: tc for tc in db.query(AssetTypeConfig).all()}
        assets = db.query(Asset).filter(Asset.is_active == True).all()
        departments = {d.code: d for d in db.query(Department).filter(Department.is_active == True).all()}
        rooms = db.query(Room).filter(Room.is_active == True).all()

        total_count = len(assets)
        total_purchase = sum(safe_decimal_to_float(a.purchase_price) for a in assets)
        total_current = sum(safe_decimal_to_float(a.current_value) for a in assets)

        # По типам
        type_breakdown = {}
        for a in assets:
            t = a.asset_type or "unknown"
            if t not in type_breakdown:
                tc = type_configs.get(t)
                type_breakdown[t] = {
                    "code": t, "name": tc.name if tc else t, "icon": tc.icon if tc else "📦",
                    "count": 0, "total_purchase": 0.0, "total_current": 0.0,
                }
            type_breakdown[t]["count"] += 1
            type_breakdown[t]["total_purchase"] += safe_decimal_to_float(a.purchase_price)
            type_breakdown[t]["total_current"] += safe_decimal_to_float(a.current_value)

        # По статусам
        status_labels = {"active": "Активен", "maintenance": "На ремонте", "reserved": "В резерве",
                         "decommissioned": "Выведен", "lost": "Утерян", "written_off": "Списан"}
        status_breakdown = {}
        for a in assets:
            s = a.status or "unknown"
            if s not in status_breakdown:
                status_breakdown[s] = {"status": s, "label": status_labels.get(s, s), "count": 0, "total_current": 0.0}
            status_breakdown[s]["count"] += 1
            status_breakdown[s]["total_current"] += safe_decimal_to_float(a.current_value)

        # По подразделениям
        dept_breakdown = {}
        for a in assets:
            dept_code = a.department_code or "Без подразделения"
            if dept_code not in dept_breakdown:
                dept = departments.get(dept_code)
                dept_breakdown[dept_code] = {
                    "department_code": dept_code,
                    "department_name": dept.name if dept else dept_code,
                    "count": 0, "total_purchase": 0.0, "total_current": 0.0,
                    "responsible_persons": set(),
                }
            dept_breakdown[dept_code]["count"] += 1
            dept_breakdown[dept_code]["total_purchase"] += safe_decimal_to_float(a.purchase_price)
            dept_breakdown[dept_code]["total_current"] += safe_decimal_to_float(a.current_value)
            if a.responsible_person:
                dept_breakdown[dept_code]["responsible_persons"].add(a.responsible_person)

        for db_item in dept_breakdown.values():
            db_item["responsible_persons"] = list(db_item["responsible_persons"])

        # По помещениям
        location_breakdown = {}
        for a in assets:
            loc = a.location_address or "Не указано"
            if loc not in location_breakdown:
                location_breakdown[loc] = {"location": loc, "count": 0, "total_current": 0.0}
            location_breakdown[loc]["count"] += 1
            location_breakdown[loc]["total_current"] += safe_decimal_to_float(a.current_value)

        # Активы без ответственного
        no_responsible = [a for a in assets if not a.responsible_person]

        # Активы, требующие ремонта
        needs_repair_count = len([a for a in assets if a.status == "maintenance"])

        return JSONResponse(content={
            "title": "Инвентаризационный отчет",
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "total_count": total_count,
                "total_purchase_value": round(total_purchase, 2),
                "total_current_value": round(total_current, 2),
                "needs_repair_count": needs_repair_count,
                "no_responsible_count": len(no_responsible),
                "department_count": len([a for a in assets if a.department_code]),
                "avg_value_per_asset": round(total_current / total_count, 2) if total_count > 0 else 0,
            },
            "type_breakdown": sorted(type_breakdown.values(), key=lambda x: x["count"], reverse=True),
            "status_breakdown": sorted(status_breakdown.values(), key=lambda x: x["count"], reverse=True),
            "department_breakdown": sorted(dept_breakdown.values(), key=lambda x: x["count"], reverse=True)[:20],
            "location_breakdown": sorted(location_breakdown.values(), key=lambda x: x["count"], reverse=True)[:20],
            "no_responsible_count": len(no_responsible),
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/inventory-report/export")
async def export_inventory_report(
    format: str = Query("excel", description="Формат: excel, pdf, json"),
    db: Session = Depends(get_db),
):
    """Экспорт инвентаризационного отчета"""
    try:
        report_data = await get_inventory_report(db)
        data = json.loads(report_data.body.decode())

        if format == "excel":
            wb = Workbook()
            ws = get_worksheet(wb)
            ws.title = "Инвентаризация"

            ws.merge_cells('A1:E1')
            ws['A1'] = data['title']
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal="center")

            row = 3
            s = data['summary']
            for label, val in [("Всего активов", s['total_count']),
                               ("Общая стоимость покупки", f"{s['total_purchase_value']:,.2f} ₽"),
                               ("Текущая стоимость", f"{s['total_current_value']:,.2f} ₽"),
                               ("Средняя стоимость", f"{s['avg_value_per_asset']:,.2f} ₽"),
                               ("Требуют ремонта", s['needs_repair_count']),
                               ("Без ответственного", s['no_responsible_count'])]:
                _style_header(ws, row, 1, label)
                _style_cell(ws, row, 2, val)
                row += 1

            row += 2
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1, value="По типам активов").font = Font(bold=True, size=14)
            row += 1
            for col, h in enumerate(["Тип", "Количество", "Стоимость", "Доля по кол-ву", "Доля по стоимости"], 1):
                _style_header(ws, row, col, h)
            row += 1
            for t in data['type_breakdown']:
                _style_cell(ws, row, 1, f"{t['icon']} {t['name']}")
                _style_cell(ws, row, 2, t['count'])
                _style_cell(ws, row, 3, f"{t['total_current']:,.2f} ₽")
                _style_cell(ws, row, 4, f"{t['count'] / s['total_count'] * 100:.1f}%" if s['total_count'] > 0 else "0%")
                _style_cell(ws, row, 5, f"{t['total_current'] / s['total_current_value'] * 100:.1f}%" if s['total_current_value'] > 0 else "0%")
                row += 1

            row += 2
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row=row, column=1, value="По статусам").font = Font(bold=True, size=14)
            row += 1
            for col, h in enumerate(["Статус", "Количество", "Стоимость", "Доля"], 1):
                _style_header(ws, row, col, h)
            row += 1
            for st in data['status_breakdown']:
                _style_cell(ws, row, 1, st['label'])
                _style_cell(ws, row, 2, st['count'])
                _style_cell(ws, row, 3, f"{st['total_current']:,.2f} ₽")
                _style_cell(ws, row, 4, f"{st['count'] / s['total_count'] * 100:.1f}%" if s['total_count'] > 0 else "0%")
                row += 1

            row += 2
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1, value="По подразделениям").font = Font(bold=True, size=14)
            row += 1
            for col, h in enumerate(["Подразделение", "Количество", "Стоимость", "Доля", "Ответственные"], 1):
                _style_header(ws, row, col, h)
            row += 1
            for d in data['department_breakdown']:
                _style_cell(ws, row, 1, d['department_name'])
                _style_cell(ws, row, 2, d['count'])
                _style_cell(ws, row, 3, f"{d['total_current']:,.2f} ₽")
                _style_cell(ws, row, 4, f"{d['count'] / s['total_count'] * 100:.1f}%" if s['total_count'] > 0 else "0%")
                _style_cell(ws, row, 5, ", ".join(d['responsible_persons'][:3]))
                row += 1

            for col_letter in ['A', 'B', 'C', 'D', 'E']:
                ws.column_dimensions[col_letter].width = 25

            filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(REPORTS_DIR, filename)
            wb.save(filepath)
        else:
            filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            filepath = os.path.join(REPORTS_DIR, filename)
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

        return FileResponse(path=filepath, filename=filename, media_type="application/octet-stream")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


# ======================================================================
# ОТЧЁТ ПО ИМПОРТУ
# ======================================================================

@router.get("/import-report")
async def get_import_report(db: Session = Depends(get_db)):
    """Отчет по импорту — реальные данные из журнала импорта"""
    try:
        from src.infrastructure.db.models.import_job import ImportJob
        from src.infrastructure.db.models.user import User

        jobs = db.query(ImportJob).order_by(ImportJob.created_at.desc()).limit(50).all()
        users = {u.id: u.username for u in db.query(User).all()}

        total_imports = db.query(ImportJob).count()
        successful_imports = db.query(ImportJob).filter(ImportJob.status == "completed").count()
        failed_imports = db.query(ImportJob).filter(ImportJob.status == "failed").count()
        total_rows_imported = db.query(func.coalesce(func.sum(ImportJob.successful_rows), 0)).scalar() or 0
        total_rows_failed = db.query(func.coalesce(func.sum(ImportJob.failed_rows), 0)).scalar() or 0

        # По типам импорта
        type_stats = db.query(
            ImportJob.import_type,
            func.count(ImportJob.id).label("count"),
            func.sum(func.coalesce(ImportJob.successful_rows, 0)).label("total_rows"),
        ).group_by(ImportJob.import_type).all()

        import_type_labels = {
            "assets": "Активы",
            "inventory": "Инвентаризация",
            "employees": "Сотрудники",
            "departments": "Подразделения",
        }

        jobs_list = []
        for job in jobs:
            jobs_list.append({
                "id": job.id,
                "filename": job.filename,
                "import_type": import_type_labels.get(job.import_type, job.import_type),
                "status": job.status,
                "total_rows": job.total_rows,
                "successful_rows": job.successful_rows,
                "failed_rows": job.failed_rows,
                "created_at": safe_isoformat(job.created_at),
                "completed_at": safe_isoformat(job.completed_at),
                "created_by": users.get(job.created_by, f"ID:{job.created_by}"),
                "summary": job.summary,
            })

        status_labels = {
            "pending": "Ожидает",
            "processing": "Выполняется",
            "completed": "Завершён",
            "failed": "Ошибка",
        }

        return JSONResponse(content={
            "title": "Отчет об импорте",
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "total_imports": total_imports,
                "successful_imports": successful_imports,
                "failed_imports": failed_imports,
                "total_rows_imported": int(total_rows_imported),
                "total_rows_failed": int(total_rows_failed),
                "success_rate": round(successful_imports / total_imports * 100, 1) if total_imports > 0 else 0,
            },
            "type_stats": [
                {
                    "type": import_type_labels.get(ts.import_type, ts.import_type),
                    "count": ts.count,
                    "total_rows": int(ts.total_rows),
                }
                for ts in type_stats
            ],
            "status_labels": status_labels,
            "recent_jobs": jobs_list[:10],
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/import-report/export")
async def export_import_report(
    format: str = Query("excel", description="Формат: excel, pdf, json"),
    db: Session = Depends(get_db),
):
    """Экспорт отчета по импорту"""
    try:
        report_data = await get_import_report(db)
        data = json.loads(report_data.body.decode())

        if format == "excel":
            wb = Workbook()
            ws = get_worksheet(wb)
            ws.title = "Импорт"

            ws.merge_cells('A1:F1')
            ws['A1'] = data['title']
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal="center")

            row = 3
            s = data['summary']
            for label, val in [("Всего импортов", s['total_imports']),
                               ("Успешных", s['successful_imports']),
                               ("С ошибками", s['failed_imports']),
                               ("Успешность", f"{s['success_rate']}%"),
                               ("Импортировано строк", s['total_rows_imported']),
                               ("Строк с ошибками", s['total_rows_failed'])]:
                _style_header(ws, row, 1, label)
                _style_cell(ws, row, 2, val)
                row += 1

            row += 2
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row=row, column=1, value="Последние импорты").font = Font(bold=True, size=14)
            row += 1
            for col, h in enumerate(["Файл", "Тип", "Статус", "Строк", "Успешно", "Ошибок", "Дата"], 1):
                _style_header(ws, row, col, h)
            row += 1
            for j in data['recent_jobs']:
                vals = [j['filename'], j['import_type'], j['status'],
                        j['total_rows'], j['successful_rows'], j['failed_rows'],
                        j['created_at'][:10] if j['created_at'] else '']
                for col, val in enumerate(vals, 1):
                    _style_cell(ws, row, col, val)
                row += 1

            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws.column_dimensions[col_letter].width = 20

            filename = f"import_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(REPORTS_DIR, filename)
            wb.save(filepath)
        else:
            data_to_save = {
                "title": data['title'],
                "generated_at": data['generated_at'],
                "summary": data['summary'],
                "type_stats": data['type_stats'],
                "recent_jobs": data['recent_jobs'],
            }
            filename = f"import_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            filepath = os.path.join(REPORTS_DIR, filename)
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data_to_save, f, ensure_ascii=False, indent=2)

        return FileResponse(path=filepath, filename=filename, media_type="application/octet-stream")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@router.get("/health")
async def reports_health():
    return JSONResponse(content={
        "status": "ok",
        "service": "reports",
        "timestamp": datetime.now().isoformat()
    })