# backend/src/presentation/http/routers/assets_export.py
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.infrastructure.db.init_db import get_db
from src.infrastructure.db.models.asset import Asset

router = APIRouter(prefix="/api/export", tags=["assets-export"])


def _format_date(value) -> str:
    if not value:
        return ""
    try:
        if hasattr(value, 'strftime'):
            return value.strftime('%d.%m.%Y')
        return str(value)
    except (AttributeError, ValueError):
        return ""


def _format_money(value) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


@router.get("/assets")
async def export_assets_excel(
    status: str = None,
    search: str = None,
    location: str = None,
    department: str = None,
    responsible: str = None,
    employee: str = None,
    active_only: bool = True,
    include_hidden: bool = False,
    db: Session = Depends(get_db),
):
    """Экспорт активов в Excel файл (.xlsx)"""
    try:
        query = db.query(Asset).options(joinedload(Asset.asset_type_config))

        if not include_hidden:
            query = query.filter(Asset.is_active == True)

        if status:
            query = query.filter(Asset.status == status)
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                (Asset.name.ilike(search_pattern)) |
                (Asset.inventory_number.ilike(search_pattern)) |
                (Asset.responsible_person.ilike(search_pattern))
            )
        if location:
            query = query.filter(Asset.location_address.ilike(f"%{location}%"))
        if department:
            query = query.filter(Asset.department_code.ilike(f"%{department}%"))
        if responsible:
            query = query.filter(Asset.responsible_person.ilike(f"%{responsible}%"))
        if employee:
            query = query.filter(Asset.responsible_person.ilike(f"%{employee}%"))

        assets = query.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Активы"

        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(vertical='center', wrap_text=True)
        date_alignment = Alignment(horizontal='center', vertical='center')
        money_alignment = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = [
            '№',
            'Инв. номер',
            'Название',
            'Тип',
            'Статус',
            'Количество',
            'Стоимость',
            'Расположение',
            'Кому выдан',
            'Дата постановки на учет',
            'Дата окончания полезного срока эксплуатации',
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        row_num = 2
        for idx, asset in enumerate(assets, 1):
            commissioning_date = getattr(asset, 'commissioning_date', None)
            depreciation_years = getattr(asset, 'depreciation_years', None)
            end_use_date = ""
            if commissioning_date and depreciation_years:
                try:
                    end_date = commissioning_date.replace(
                        year=commissioning_date.year + depreciation_years
                    )
                    end_use_date = end_date.strftime('%d.%m.%Y')
                except (ValueError, TypeError):
                    end_use_date = ""

            employee_name = ""
            if getattr(asset, 'employee_name', None):
                employee_name = asset.employee_name
            elif getattr(asset, 'responsible_person', None):
                employee_name = asset.responsible_person

            row_data = [
                idx,
                getattr(asset, 'inventory_number', ''),
                getattr(asset, 'name', ''),
                getattr(asset, 'asset_type', ''),
                getattr(asset, 'status', ''),
                getattr(asset, 'quantity', 1),
                _format_money(getattr(asset, 'purchase_price', None)),
                getattr(asset, 'location_address', '') or '',
                employee_name,
                _format_date(commissioning_date),
                end_use_date,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = cell_alignment

                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == 5:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == 7:
                    cell.number_format = '#,##0.00'
                    cell.alignment = money_alignment
                elif col_idx in (10, 11):
                    cell.alignment = date_alignment

            row_num += 1

        for col_idx in range(1, len(headers) + 1):
            max_length = len(str(headers[col_idx - 1]))
            for row in range(2, row_num):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    cell_len = len(str(cell_value))
                    if cell_len > max_length:
                        max_length = cell_len
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 40)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"assets_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            iter([output.read()]),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
