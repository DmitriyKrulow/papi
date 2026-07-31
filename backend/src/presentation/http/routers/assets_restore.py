# backend/src/presentation/http/routers/assets_restore.py
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from typing import List

from src.infrastructure.db.init_db import get_db
from src.infrastructure.db.models.asset import Asset
from src.infrastructure.db.models.repair_request import RepairRequest
from src.infrastructure.db.models.maintenance_event import MaintenanceEvent
from src.infrastructure.db.models.maintenance_record import MaintenanceRecord
from src.infrastructure.db.models.movement_record import MovementRecord
from src.infrastructure.db.models.depreciation_record import DepreciationRecord
from src.infrastructure.db.models.asset_photo import AssetPhoto
from src.infrastructure.db.models.user import User
from src.presentation.http.dependencies.auth import get_current_admin

router = APIRouter(prefix="/api/admin/assets", tags=["assets-restore"])


@router.post("/restore-from-excel")
async def restore_assets_from_excel(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    """
    Восстановить базу активов из Excel файла.
    Удаляет все текущие активы и создаёт новые из файла.
    Формат файла должен совпадать с форматом экспорта.
    """
    try:
        # Читаем сырое тело запроса как байты
        body = await request.body()
        if not body:
            raise HTTPException(status_code=400, detail="Файл не передан")
        
        wb = load_workbook(filename=BytesIO(body))
        ws = wb.active
        
        # Читаем заголовки из первой строки
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else '')
        
        # Пропускаем пустые строки в начале
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and any(v is not None for v in row):
                data_rows.append(row)
        
        if not data_rows:
            raise HTTPException(status_code=400, detail="Файл не содержит данных")
        
        # Удаляем все связанные данные и текущие активы
        # Порядок важен: сначала дочерние таблицы (внешние ключи)
        db.query(RepairRequest).delete()
        db.query(MaintenanceEvent).delete()
        db.query(MaintenanceRecord).delete()
        db.query(MovementRecord).delete()
        db.query(DepreciationRecord).delete()
        db.query(AssetPhoto).delete()
        # Теперь сами активы
        deleted_count = db.query(Asset).delete()
        db.commit()
        
        # Создаём новые активы из файла
        created_count = 0
        errors = []
        
        for idx, row in enumerate(data_rows, start=2):
            try:
                # Создаём словарь из заголовков и данных строки
                row_data = {}
                for h, v in zip(headers, row):
                    row_data[h] = v
                
                # Ищем нужные колонки по разным возможным названиям
                inv_number = str(row_data.get('Инв. номер', '') or row_data.get('Инв.номер', '') or '')
                name = str(row_data.get('Название', '') or '')
                
                if not inv_number or not name:
                    errors.append(f"Строка {idx}: пропущена (нет инв. номера или названия)")
                    continue
                
                # Проверяем уникальность
                existing = db.query(Asset).filter(
                    Asset.inventory_number == inv_number
                ).first()
                
                if existing:
                    errors.append(f"Строка {idx}: пропущен дубликат инв. номера {inv_number}")
                    continue
                
                # Парсим данные
                purchase_price = None
                val = row_data.get('Стоимость', 0)
                if val:
                    try:
                        purchase_price = Decimal(str(float(val)))
                    except:
                        pass
                
                quantity = 1
                val = row_data.get('Количество', 1)
                if val:
                    try:
                        quantity = int(float(val))
                    except:
                        pass
                
                asset_status = 'active'
                status_val = row_data.get('Статус', '')
                if status_val:
                    status_map = {
                        'активен': 'active',
                        'active': 'active',
                        'на ремонте': 'maintenance',
                        'maintenance': 'maintenance',
                        'в резерве': 'reserved',
                        'reserved': 'reserved',
                        'списан': 'written_off',
                        'written_off': 'written_off',
                        'утерян': 'lost',
                        'lost': 'lost',
                    }
                    asset_status = status_map.get(str(status_val).strip().lower(), 'active')
                
                asset = Asset(
                    inventory_number=inv_number,
                    name=name,
                    description=str(row_data.get('Описание', '') or '') or None,
                    model=str(row_data.get('Модель', '') or '') or None,
                    asset_type=str(row_data.get('Тип', '') or '') or None,
                    status=asset_status,
                    purchase_price=purchase_price,
                    current_value=purchase_price,
                    quantity=quantity,
                    department_code=str(row_data.get('Подразделение', '') or '') or None,
                    responsible_person=str(row_data.get('Кому выдан', '') or '') or None,
                    location_address=str(row_data.get('Расположение', '') or '') or None,
                    manufacturer_code=str(row_data.get('Код производителя', '') or '') or None,
                    manufacturer_name=str(row_data.get('Производитель', '') or '') or None,
                    commissioning_date=None,
                    is_active=True,
                    created_at=datetime.now(),
                    updated_at=datetime.now(),
                )
                
                db.add(asset)
                created_count += 1
                
            except Exception as e:
                errors.append(f"Строка {idx}: ошибка - {str(e)}")
                continue
        
        db.commit()
        
        return {
            "message": f"База восстановлена: удалено {deleted_count}, создано {created_count}",
            "deleted": deleted_count,
            "created": created_count,
            "errors": errors[:10],
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
