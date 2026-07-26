# src/infrastructure/reports/repair_report.py
from typing import Dict, List, Optional
from datetime import datetime

import pandas as pd

from .base import BaseReportGenerator


class RepairReportGenerator(BaseReportGenerator):
    """Генератор отчета по заявкам на ремонт"""

    def __init__(self):
        super().__init__()
        self.title = "ЗАЯВКА НА РЕМОНТ"
        self.date = datetime.now().strftime('%d.%m.%Y')

    def generate(
        self,
        repairs: List[Dict],
        template_data: Dict = None,
        **kwargs
    ) -> bytes:
        """Генерация отчета по заявкам на ремонт"""
        
        df = pd.DataFrame(repairs)
        
        if df.empty:
            df = pd.DataFrame(columns=[
                'ID', 'Название', 'Описание', 'Приоритет', 'Статус',
                'Актив', 'Инвентарный номер', 'Создано', 'Исполнитель', 'Стоимость'
            ])
        
        return self.save_bytes()

    def generate_single_repair(self, repair_data: Dict) -> bytes:
        """Генерация PDF для одной заявки на ремонт"""
        
        sheet = self.workbook.create_sheet("Заявка на ремонт")
        
        # Заголовок
        sheet.merge_cells('A1:D1')
        title_cell = sheet['A1']
        title_cell.value = self.title
        title_cell.font = self._get_title_font()
        title_cell.alignment = self._get_center_alignment()
        
        sheet['A3'] = f"Дата генерации: {self.date}"
        
        # Информация о заявке
        row = 5
        sheet[f'A{row}'] = "Номер заявки:"
        sheet[f'B{row}'] = repair_data.get('id', '—')
        row += 1
        
        sheet[f'A{row}'] = "Дата создания:"
        sheet[f'B{row}'] = repair_data.get('created_at', '—')[:10] if repair_data.get('created_at') else '—'
        row += 1
        
        sheet[f'A{row}'] = "Статус:"
        sheet[f'B{row}'] = repair_data.get('status', '—')
        row += 1
        
        sheet[f'A{row}'] = "Приоритет:"
        sheet[f'B{row}'] = repair_data.get('priority', '—')
        row += 2
        
        # Информация об активе
        sheet[f'A{row}'] = "ИНФОРМАЦИЯ ОБ АКТИВЕ"
        sheet[f'A{row}'].font = self._get_header_font()
        row += 1
        
        sheet[f'A{row}'] = "Наименование:"
        sheet[f'B{row}'] = repair_data.get('asset_name', '—')
        row += 1
        
        sheet[f'A{row}'] = "Инвентарный номер:"
        sheet[f'B{row}'] = repair_data.get('inventory_number', '—')
        row += 1
        
        sheet[f'A{row}'] = "Ответственный:"
        sheet[f'B{row}'] = repair_data.get('responsible_person', '—')
        row += 2
        
        # Описание работ
        sheet[f'A{row}'] = "ОПИСАНИЕ РАБОТ"
        sheet[f'A{row}'].font = self._get_header_font()
        row += 1
        
        description = repair_data.get('description', '—')
        sheet[f'A{row}'] = description
        sheet[f'A{row}'].alignment = self._get_wrap_alignment()
        row += 3
        
        # Дополнительная информация
        sheet[f'A{row}'] = "ДОПОЛНИТЕЛЬНАЯ ИНФОРМАЦИЯ"
        sheet[f'A{row}'].font = self._get_header_font()
        row += 1
        
        sheet[f'A{row}'] = "Желаемая дата выполнения:"
        sheet[f'B{row}'] = repair_data.get('desired_completion_date', '—')
        row += 1
        
        sheet[f'A{row}'] = "Ориентировочная стоимость:"
        sheet[f'B{row}'] = f"{repair_data.get('estimated_cost', 0)} ₽"
        row += 1
        
        sheet[f'A{row}'] = "Исполнитель:"
        sheet[f'B{row}'] = repair_data.get('assigned_to_name', 'Не назначен')
        row += 2
        
        # Подписи
        sheet[f'A{row}'] = "_________________________"
        row += 1
        sheet[f'A{row}'] = "Подпись ответственного лица"
        row += 2
        sheet[f'A{row}'] = "_________________________"
        row += 1
        sheet[f'A{row}'] = "Подпись исполнителя"
        
        # Адаптация ширины
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 40
        
        return self.save_bytes()

    def _get_title_font(self):
        from openpyxl.styles import Font
        return Font(size=16, bold=True)

    def _get_header_font(self):
        from openpyxl.styles import Font
        return Font(bold=True)

    def _get_header_fill(self):
        from openpyxl.styles import PatternFill
        return PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    def _get_center_alignment(self):
        from openpyxl.styles import Alignment
        return Alignment(horizontal='center')

    def _get_right_alignment(self):
        from openpyxl.styles import Alignment
        return Alignment(horizontal='right')

    def _get_wrap_alignment(self):
        from openpyxl.styles import Alignment
        return Alignment(wrap_text=True, vertical='top')

    def _get_border(self):
        from openpyxl.styles import Border, Side
        thin = Side(border_style='thin', color='000000')
        return Border(left=thin, right=thin, top=thin, bottom=thin)
