# src/infrastructure/reports/base.py
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Dict, List, Optional
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class BaseReportGenerator(ABC):
    """Базовый класс для генерации отчетов"""

    def __init__(self):
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

    @abstractmethod
    def generate(self, data: Dict[str, Any], **kwargs) -> bytes:
        """Генерация отчета"""
        pass

    def _add_sheet(self, name: str, df: pd.DataFrame, styles: Optional[Dict] = None) -> None:
        """Добавление листа с таблицей"""
        sheet = self.workbook.create_sheet(name)
        
        if not df.empty:
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    self._apply_cell_styles(cell, styles)

    def _apply_cell_styles(self, cell, styles: Optional[Dict] = None) -> None:
        """Применение стилей к ячейке"""
        if styles:
            if styles.get('bold'):
                cell.font = Font(bold=True)
            if styles.get('bg_color'):
                cell.fill = PatternFill(start_color=styles['bg_color'], end_color=styles['bg_color'], fill_type='solid')
            if styles.get('center'):
                cell.alignment = Alignment(horizontal='center')
            if styles.get('right'):
                cell.alignment = Alignment(horizontal='right')
            if styles.get('wrap'):
                cell.alignment = Alignment(wrap_text=True)
            
            thin = Side(border_style='thin', color='000000')
            if styles.get('border'):
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _format_currency(self, value: float, decimals: int = 2) -> str:
        """Форматирование валюты"""
        if value is None or pd.isna(value):
            return '0.00'
        return f"{value:,.{decimals}f}".replace(',', ' ')

    def _format_date(self, value: Any) -> str:
        """Форматирование даты"""
        if value is None or pd.isna(value):
            return ''
        if isinstance(value, datetime):
            return value.strftime('%d.%m.%Y')
        return str(value)

    def _add_summary_row(self, sheet, row: int, values: List[str], styles: Optional[Dict] = None) -> None:
        """Добавление строки итогов"""
        for col_idx, value in enumerate(values, 1):
            cell = sheet.cell(row=row, column=col_idx, value=value)
            self._apply_cell_styles(cell, {'bold': True, 'bg_color': 'E0E0E0', 'border': True})

    def save(self, path: str) -> None:
        """Сохранение в файл"""
        self.workbook.save(path)

    def save_bytes(self) -> bytes:
        """Сохранение в байты"""
        from io import BytesIO
        output = BytesIO()
        self.workbook.save(output)
        return output.getvalue()

    def get_sheet_names(self) -> List[str]:
        """Получение имен листов"""
        return self.workbook.sheetnames
