# src/infrastructure/reports/depreciation_report.py
from typing import Dict, List, Optional
from datetime import datetime

import pandas as pd

from .base import BaseReportGenerator


class DepreciationReportGenerator(BaseReportGenerator):
    """Генератор отчета по амортизации"""

    def __init__(self):
        super().__init__()
        self.title = "Отчет по амортизации"
        self.date = datetime.now().strftime('%d.%m.%Y')

    def generate(
        self,
        assets: List[Dict],
        depreciation_records: List[Dict] = None,
        report_date: datetime = None,
        **kwargs
    ) -> bytes:
        """Генерация отчета по амортизации"""
        
        df = pd.DataFrame(assets)
        
        if df.empty:
            df = pd.DataFrame(columns=[
                'ID', 'Инвентарный номер', 'Название', 'Стоимость', 'Срок службы',
                'Год начала', 'Год окончания', 'Ежемесячная амортизация',
                'Накопленная амортизация', 'Остаточная стоимость', 'Статус'
            ])
        
        self._add_summary_sheet(df)
        self._add_depreciation_sheet(df, depreciation_records or [])
        self._add_by_category_sheet(df)
        
        return self.save_bytes()

    def _add_summary_sheet(self, df: pd.DataFrame) -> None:
        """Добавление листа сводки"""
        sheet = self.workbook.create_sheet("Сводка")
        
        sheet.merge_cells('A1:D1')
        title_cell = sheet['A1']
        title_cell.value = self.title
        title_cell.font = self._get_title_font()
        title_cell.alignment = self._get_center_alignment()
        
        sheet['A3'] = f"Дата генерации: {self.date}"
        sheet['A4'] = f"Всего активов: {len(df)}"
        
        if not df.empty:
            total_purchase = df['purchase_price'].sum() if 'purchase_price' in df.columns else 0
            total_residual = df['residual_value'].sum() if 'residual_value' in df.columns else 0
            
            sheet['A5'] = f"Первоначальная стоимость: {self._format_currency(total_purchase)}"
            sheet['A6'] = f"Остаточная стоимость: {self._format_currency(total_residual)}"
            
            if 'depreciation_rate' in df.columns:
                avg_rate = df['depreciation_rate'].mean()
                sheet['A7'] = f"Средняя ставка амортизации: {avg_rate:.2f}%"
        
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20

    def _add_depreciation_sheet(self, df: pd.DataFrame, records: List[Dict]) -> None:
        """Добавление листа амортизации"""
        sheet = self.workbook.create_sheet("Амортизация")
        
        sheet.merge_cells('A1:D1')
        title_cell = sheet['A1']
        title_cell.value = "Амортизация по активам"
        title_cell.font = self._get_title_font()
        title_cell.alignment = self._get_center_alignment()
        
        headers = [
            'ID', 'Инвентарный номер', 'Название', 'Стоимость', 'Срок службы',
            'Ставка', 'Ежемесячная', 'Накопленная', 'Остаточная', 'Статус'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col_idx, value=header)
            cell.font = self._get_header_font()
            cell.fill = self._get_header_fill()
            cell.border = self._get_border()
        
        for row_idx, (_, row) in enumerate(df.iterrows(), 4):
            purchase_price = row.get('purchase_price', 0) or 0
            residual_value = row.get('residual_value', 0) or 0
            depreciation_rate = row.get('depreciation_rate', 0) or 0
            useful_life = row.get('useful_life', 0) or 0
            
            monthly_depreciation = purchase_price * (depreciation_rate / 100 / 12) if useful_life > 0 else 0
            accumulated = monthly_depreciation * useful_life if useful_life > 0 else 0
            book_value = purchase_price - accumulated
            
            sheet.cell(row=row_idx, column=1, value=row.get('id', ''))
            sheet.cell(row=row_idx, column=2, value=row.get('inventory_number', ''))
            sheet.cell(row=row_idx, column=3, value=row.get('name', ''))
            sheet.cell(row=row_idx, column=4, value=self._format_currency(purchase_price))
            sheet.cell(row=row_idx, column=5, value=str(useful_life) if useful_life else '')
            sheet.cell(row=row_idx, column=6, value=f"{depreciation_rate:.2f}%" if depreciation_rate else '')
            sheet.cell(row=row_idx, column=7, value=self._format_currency(monthly_depreciation))
            sheet.cell(row=row_idx, column=8, value=self._format_currency(accumulated))
            sheet.cell(row=row_idx, column=9, value=self._format_currency(book_value))
            sheet.cell(row=row_idx, column=10, value=row.get('status', ''))
        
        total_row = 4 + len(df)
        if not df.empty:
            sheet.cell(row=total_row, column=1, value="Итого:")
            sheet.cell(row=total_row, column=4, value=self._format_currency(df['purchase_price'].sum() if 'purchase_price' in df.columns else 0))
            sheet.cell(row=total_row, column=9, value=self._format_currency(df['residual_value'].sum() if 'residual_value' in df.columns else 0))
            
            for col in range(1, 11):
                cell = sheet.cell(row=total_row, column=col)
                cell.font = self._get_header_font()
                cell.fill = self._get_header_fill()
                cell.border = self._get_border()
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            sheet.column_dimensions[col].width = 18

    def _add_by_category_sheet(self, df: pd.DataFrame) -> None:
        """Добавление листа по категориям"""
        sheet = self.workbook.create_sheet("По категориям")
        
        sheet.merge_cells('A1:C1')
        title_cell = sheet['A1']
        title_cell.value = "Амортизация по категориям"
        title_cell.font = self._get_title_font()
        title_cell.alignment = self._get_center_alignment()
        
        if not df.empty:
            # Группировка по типу актива
            if 'asset_type' in df.columns:
                groups = df.groupby('asset_type').agg({
                    'id': 'count',
                    'purchase_price': 'sum',
                    'residual_value': 'sum'
                }).reset_index()
                
                groups.columns = ['Категория', 'Количество', 'Первоначальная', 'Остаточная']
                
                headers = groups.columns.tolist()
                for col_idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=3, column=col_idx, value=header)
                    cell.font = self._get_header_font()
                    cell.fill = self._get_header_fill()
                    cell.border = self._get_border()
                
                for row_idx, (_, row) in enumerate(groups.iterrows(), 4):
                    for col_idx, value in enumerate(row, 1):
                        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = self._get_border()
                        if col_idx > 1:
                            cell.alignment = self._get_right_alignment()
                
                # Итого
                total_row = 4 + len(groups)
                sheet.cell(row=total_row, column=1, value="Итого:")
                sheet.cell(row=total_row, column=2, value=len(df))
                
                if 'purchase_price' in df.columns:
                    sheet.cell(row=total_row, column=3, value=self._format_currency(df['purchase_price'].sum()))
                    sheet.cell(row=total_row, column=4, value=self._format_currency(df['residual_value'].sum() if 'residual_value' in df.columns else 0))
                
                for col in range(1, 5):
                    cell = sheet.cell(row=total_row, column=col)
                    cell.font = self._get_header_font()
                    cell.fill = self._get_header_fill()
                    cell.border = self._get_border()

    def _get_title_font(self):
        from openpyxl.styles import Font
        return Font(size=16, bold=True)

    def _get_header_font(self):
        from openpyxl.styles import Font
        return Font(bold=True)

    def _get_header_fill(self):
        from openpyxl.styles import PatternFill
        return PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    def _get_center_alignment(self):
        from openpyxl.styles import Alignment
        return Alignment(horizontal='center')

    def _get_right_alignment(self):
        from openpyxl.styles import Alignment
        return Alignment(horizontal='right')

    def _get_border(self):
        from openpyxl.styles import Border, Side
        thin = Side(border_style='thin', color='000000')
        return Border(left=thin, right=thin, top=thin, bottom=thin)
