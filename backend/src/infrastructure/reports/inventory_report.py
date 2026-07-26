# backend/src/infrastructure/reports/inventory_report.py
from typing import Dict, List, Optional
from datetime import datetime

import pandas as pd

from .base import BaseReportGenerator


class InventoryReportGenerator(BaseReportGenerator):
    """Генератор инвентаризационной ведомости"""

    def __init__(self):
        super().__init__()
        self.title = "ИНВЕНТАРИЗАЦИОННАЯ ВЕДОМОСТЬ"
        self.date = datetime.now().strftime('%d.%m.%Y')

    def generate(
        self,
        assets: List[Dict],
        department_name: str = "",
        **kwargs
    ) -> bytes:
        """Генерация инвентаризационной ведомости"""
        
        df = pd.DataFrame(assets)
        
        if df.empty:
            df = pd.DataFrame(columns=[
                'ID', 'Инвентарный номер', 'Наименование', 'Модель',
                'Производитель', 'Состояние', 'Местоположение', 'Ответственный'
            ])
        
        self._add_inventory_sheet(df, department_name)
        
        return self.save_bytes()

    def _add_inventory_sheet(self, df: pd.DataFrame, department_name: str) -> None:
        """Добавление листа инвентаризации"""
        sheet = self.workbook.create_sheet("Инвентаризация")
        
        # Заголовок
        sheet.merge_cells('A1:H1')
        title_cell = sheet['A1']
        title_cell.value = self.title
        title_cell.font = self._get_title_font()
        title_cell.alignment = self._get_center_alignment()
        
        sheet['A3'] = f"Дата: {self.date}"
        if department_name:
            sheet['A4'] = f"Подразделение: {department_name}"
        sheet['A5'] = f"Всего активов: {len(df)}"
        
        # Заголовки колонок
        headers = [
            'П/п', 'Инвентарный номер', 'Наименование', 'Модель',
            'Производитель', 'Состояние', 'Местоположение', 'Ответственный'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=7, column=col_idx, value=header)
            cell.font = self._get_header_font()
            cell.fill = self._get_header_fill()
            cell.border = self._get_border()
            cell.alignment = self._get_center_alignment()
        
        # Данные
        for row_idx, (_, row) in enumerate(df.iterrows(), 8):
            sheet.cell(row=row_idx, column=1, value=row_idx - 7)
            sheet.cell(row=row_idx, column=2, value=row.get('inventory_number', '—'))
            sheet.cell(row=row_idx, column=3, value=row.get('name', '—'))
            sheet.cell(row=row_idx, column=4, value=row.get('model', '—'))
            sheet.cell(row=row_idx, column=5, value=row.get('manufacturer_name', '—'))
            sheet.cell(row=row_idx, column=6, value=row.get('status', '—'))
            sheet.cell(row=row_idx, column=7, value=row.get('location_address', '—'))
            sheet.cell(row=row_idx, column=8, value=row.get('responsible_person', '—'))
        
        # Итоговая строка
        total_row = 8 + len(df)
        for col in range(1, 9):
            cell = sheet.cell(row=total_row, column=col)
            cell.font = self._get_header_font()
            cell.fill = self._get_header_fill()
            cell.border = self._get_border()
        
        # Адаптация ширины
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 25
        sheet.column_dimensions['H'].width = 25

    def _get_title_font(self):
        from openpyxl.styles import Font
        return Font(size=16, bold=True)

    def _get_header_font(self):
        from openpyxl.styles import Font
        return Font(bold=True)

    def _get_header_fill(self):
        from openpyxl.styles import PatternFill
        return PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    def _get_center_alignment(self):
        from openpyxl.styles import Alignment
        return Alignment(horizontal='center')

    def _get_right_alignment(self):
        from openpyxl.styles import Alignment
        return Alignment(horizontal='right')

    def _get_wrap_alignment(self):
        from openpyxl.styles import Alignment
        return Alignment(wrap_text=True, vertical='top')

    def _get_border(self):
        from openpyxl.styles import Border, Side
        thin = Side(border_style='thin', color='000000')
        return Border(left=thin, right=thin, top=thin, bottom=thin)
