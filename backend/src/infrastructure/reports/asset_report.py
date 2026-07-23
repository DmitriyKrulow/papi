# src/infrastructure/reports/asset_report.py
from typing import Dict, List, Optional
from datetime import datetime

import pandas as pd

from .base import BaseReportGenerator


class AssetReportGenerator(BaseReportGenerator):
    """Генератор отчета по активам"""

    def __init__(self):
        super().__init__()
        self.title = "Отчет по активам"
        self.date = datetime.now().strftime('%d.%m.%Y')

    def generate(
        self,
        assets: List[Dict],
        include_photos: bool = False,
        include_history: bool = False,
        **kwargs
    ) -> bytes:
        """Генерация отчета по активам"""
        
        # Преобразование данных в DataFrame
        df = pd.DataFrame(assets)
        
        if df.empty:
            df = pd.DataFrame(columns=[
                'ID', 'Инвентарный номер', 'Название', 'Модель', 'Производитель',
                'Стоимость', 'Текущая стоимость', 'Статус', 'Подразделение',
                'Ответственное лицо', 'Местоположение', 'Дата покупки', 'Дата ввода',
                'Гарантия', 'Теги'
            ])
        
        # Добавление листа "Сводка"
        self._add_summary_sheet(df)
        
        # Добавление листа "Активы"
        self._add_assets_sheet(df)
        
        # Добавление листа "По подразделениям" если есть данные
        if 'department_code' in df.columns:
            self._add_by_department_sheet(df)
        
        # Добавление листа "По статусам"
        if 'status' in df.columns:
            self._add_by_status_sheet(df)
        
        return self.save_bytes()

    def _add_summary_sheet(self, df: pd.DataFrame) -> None:
        """Добавление листа сводки"""
        sheet = self.workbook.create_sheet("Сводка")
        
        # Заголовок
        sheet.merge_cells('A1:D1')
        title_cell = sheet['A1']
        title_cell.value = self.title
        title_cell.font = self._get_title_font()
        title_cell.alignment = self._get_center_alignment()
        
        sheet['A3'] = f"Дата генерации: {self.date}"
        sheet['A4'] = f"Всего активов: {len(df)}"
        
        if not df.empty:
            total_value = df['purchase_price'].sum() if 'purchase_price' in df.columns else 0
            current_value = df['current_value'].sum() if 'current_value' in df.columns else 0
            
            sheet['A5'] = f"Первоначальная стоимость: {self._format_currency(total_value)}"
            sheet['A6'] = f"Текущая стоимость: {self._format_currency(current_value)}"
            
            # Статистика по статусам
            if 'status' in df.columns:
                sheet['A8'] = "По статусам:"
                for i, (status, count) in enumerate(df['status'].value_counts().items(), 9):
                    sheet[f'A{i}'] = f"  {status}: {count}"
                    sheet[f'B{i}'] = f"{(count / len(df) * 100):.1f}%"
        
        # Адаптация ширины колонок
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20

    def _add_assets_sheet(self, df: pd.DataFrame) -> None:
        """Добавление листа с активами"""
        sheet = self.workbook.create_sheet("Активы")
        
        # Заголовок
        sheet.merge_cells('A1:D1')
        title_cell = sheet['A1']
        title_cell.value = self.title
        title_cell.font = self._get_title_font()
        title_cell.alignment = self._get_center_alignment()
        
        # Заголовки колонок
        headers = [
            'ID', 'Инвентарный номер', 'Название', 'Модель', 'Производитель',
            'Стоимость', 'Текущая стоимость', 'Статус', 'Подразделение',
            'Ответственное лицо', 'Местоположение', 'Дата покупки', 'Дата ввода'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col_idx, value=header)
            cell.font = self._get_header_font()
            cell.fill = self._get_header_fill()
            cell.alignment = self._get_center_alignment()
            cell.border = self._get_border()
        
        # Данные
        for row_idx, (_, row) in enumerate(df.iterrows(), 4):
            sheet.cell(row=row_idx, column=1, value=row.get('id', ''))
            sheet.cell(row=row_idx, column=2, value=row.get('inventory_number', ''))
            sheet.cell(row=row_idx, column=3, value=row.get('name', ''))
            sheet.cell(row=row_idx, column=4, value=row.get('model', ''))
            sheet.cell(row=row_idx, column=5, value=row.get('manufacturer_name', ''))
            sheet.cell(row=row_idx, column=6, value=self._format_currency(row.get('purchase_price', 0)))
            sheet.cell(row=row_idx, column=7, value=self._format_currency(row.get('current_value', 0)))
            sheet.cell(row=row_idx, column=8, value=row.get('status', ''))
            sheet.cell(row=row_idx, column=9, value=row.get('department_code', ''))
            sheet.cell(row=row_idx, column=10, value=row.get('responsible_person', ''))
            sheet.cell(row=row_idx, column=11, value=row.get('location_address', ''))
            sheet.cell(row=row_idx, column=12, value=self._format_date(row.get('purchase_date', '')))
            sheet.cell(row=row_idx, column=13, value=self._format_date(row.get('commissioning_date', '')))
        
        # Адаптация ширины
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            sheet.column_dimensions[col].width = 20

    def _add_by_department_sheet(self, df: pd.DataFrame) -> None:
        """Добавление листа по подразделениям"""
        sheet = self.workbook.create_sheet("По подразделениям")
        
        sheet.merge_cells('A1:D1')
        title_cell = sheet['A1']
        title_cell.value = "Активы по подразделениям"
        title_cell.font = self._get_title_font()
        title_cell.alignment = self._get_center_alignment()
        
        if not df.empty and 'department_code' in df.columns:
            dept_groups = df.groupby('department_code').agg({
                'id': 'count',
                'purchase_price': 'sum',
                'current_value': 'sum'
            }).reset_index()
            
            dept_groups.columns = ['Подразделение', 'Количество', 'Первоначальная стоимость', 'Текущая стоимость']
            
            headers = dept_groups.columns.tolist()
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=3, column=col_idx, value=header)
                cell.font = self._get_header_font()
                cell.fill = self._get_header_fill()
                cell.border = self._get_border()
            
            for row_idx, (_, row) in enumerate(dept_groups.iterrows(), 4):
                for col_idx, value in enumerate(row, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self._get_border()
                    if col_idx > 1:
                        cell.alignment = self._get_right_alignment()
            
            # Итоговая строка
            total_row = 4 + len(dept_groups)
            sheet.cell(row=total_row, column=1, value="Итого:")
            sheet.cell(row=total_row, column=2, value=len(df))
            sheet.cell(row=total_row, column=3, value=self._format_currency(df['purchase_price'].sum() if 'purchase_price' in df.columns else 0))
            sheet.cell(row=total_row, column=4, value=self._format_currency(df['current_value'].sum() if 'current_value' in df.columns else 0))
            
            for col in range(1, 5):
                cell = sheet.cell(row=total_row, column=col)
                cell.font = self._get_header_font()
                cell.fill = self._get_header_fill()
                cell.border = self._get_border()

    def _add_by_status_sheet(self, df: pd.DataFrame) -> None:
        """Добавление листа по статусам"""
        sheet = self.workbook.create_sheet("По статусам")
        
        sheet.merge_cells('A1:D1')
        title_cell = sheet['A1']
        title_cell.value = "Активы по статусам"
        title_cell.font = self._get_title_font()
        title_cell.alignment = self._get_center_alignment()
        
        if not df.empty and 'status' in df.columns:
            status_groups = df.groupby('status').agg({
                'id': 'count',
                'purchase_price': 'sum',
                'current_value': 'sum'
            }).reset_index()
            
            status_groups.columns = ['Статус', 'Количество', 'Первоначальная стоимость', 'Текущая стоимость']
            
            headers = status_groups.columns.tolist()
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=3, column=col_idx, value=header)
                cell.font = self._get_header_font()
                cell.fill = self._get_header_fill()
                cell.border = self._get_border()
            
            for row_idx, (_, row) in enumerate(status_groups.iterrows(), 4):
                for col_idx, value in enumerate(row, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self._get_border()
                    if col_idx > 1:
                        cell.alignment = self._get_right_alignment()
            
            # Итоговая строка
            total_row = 4 + len(status_groups)
            sheet.cell(row=total_row, column=1, value="Итого:")
            sheet.cell(row=total_row, column=2, value=len(df))
            sheet.cell(row=total_row, column=3, value=self._format_currency(df['purchase_price'].sum() if 'purchase_price' in df.columns else 0))
            sheet.cell(row=total_row, column=4, value=self._format_currency(df['current_value'].sum() if 'current_value' in df.columns else 0))
            
            for col in range(1, 5):
                cell = sheet.cell(row=total_row, column=col)
                cell.font = self._get_header_font()
                cell.fill = self._get_header_fill()
                cell.border = self._get_border()

    def _get_title_font(self):
        from openpyxl.styles import Font
        return Font(size=16, bold=True)

    def _get_header_font(self):
        from openpyxl.styles import Font
        return Font(bold=True)

    def _get_header_fill(self):
        from openpyxl.styles import PatternFill
        return PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    def _get_center_alignment(self):
        from openpyxl.styles import Alignment
        return Alignment(horizontal='center')

    def _get_right_alignment(self):
        from openpyxl.styles import Alignment
        return Alignment(horizontal='right')

    def _get_border(self):
        from openpyxl.styles import Border, Side
        thin = Side(border_style='thin', color='000000')
        return Border(left=thin, right=thin, top=thin, bottom=thin)
