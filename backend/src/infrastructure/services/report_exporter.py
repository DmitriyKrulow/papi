# backend/src/infrastructure/services/report_exporter.py
import os
from datetime import datetime
from typing import Dict, Any, List, Optional
from decimal import Decimal
import json

# Для Excel
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Для PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Путь для сохранения отчетов
REPORTS_DIR = "uploads/reports"
os.makedirs(REPORTS_DIR, exist_ok=True)


class ReportExporter:
    """Сервис для экспорта отчетов в различные форматы"""
    
    @staticmethod
    def export_asset_report(data: Dict[str, Any], format: str = "excel") -> str:
        """Экспортирует отчет по активам"""
        if format == "excel":
            return ReportExporter._export_asset_excel(data)
        elif format == "pdf":
            return ReportExporter._export_asset_pdf(data)
        else:
            return ReportExporter._export_asset_json(data)
    
    @staticmethod
    def export_depreciation_report(data: Dict[str, Any], format: str = "excel") -> str:
        """Экспортирует отчет по амортизации"""
        if format == "excel":
            return ReportExporter._export_depreciation_excel(data)
        elif format == "pdf":
            return ReportExporter._export_depreciation_pdf(data)
        else:
            return ReportExporter._export_depreciation_json(data)
    
    @staticmethod
    def export_inventory_report(data: Dict[str, Any], format: str = "excel") -> str:
        """Экспортирует инвентаризационный отчет"""
        if format == "excel":
            return ReportExporter._export_inventory_excel(data)
        elif format == "pdf":
            return ReportExporter._export_inventory_pdf(data)
        else:
            return ReportExporter._export_inventory_json(data)
    
    @staticmethod
    def _get_worksheet(wb: Workbook) -> Worksheet:
        """Получает активный лист или создает новый"""
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Sheet1")
        return ws
    
    @staticmethod
    def _export_asset_excel(data: Dict[str, Any]) -> str:
        """Экспорт отчета по активам в Excel"""
        wb = Workbook()
        ws = ReportExporter._get_worksheet(wb)
        ws.title = "Активы"
        
        # Стили
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = data.get('title', 'Отчет по активам')
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Дата генерации
        ws.merge_cells('A2:K2')
        ws['A2'] = f"Дата генерации: {data.get('generated_at', datetime.now().isoformat())}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Всего активов
        ws.merge_cells('A3:K3')
        ws['A3'] = f"Всего активов: {data.get('total_assets', 0)}"
        ws['A3'].alignment = Alignment(horizontal="center")
        
        # Заголовки таблицы
        headers = ['№', 'Инв. номер', 'Наименование', 'Модель', 'Тип', 'Статус', 
                   'Стоимость покупки', 'Текущая стоимость', 'Подразделение', 'Ответственный']
        row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Данные
        for idx, asset in enumerate(data.get('assets', []), 1):
            row += 1
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=asset.get('inventory_number', '')).border = border
            ws.cell(row=row, column=3, value=asset.get('name', '')).border = border
            ws.cell(row=row, column=4, value=asset.get('model', '')).border = border
            ws.cell(row=row, column=5, value=asset.get('asset_type', '')).border = border
            ws.cell(row=row, column=6, value=asset.get('status', '')).border = border
            ws.cell(row=row, column=7, value=float(asset.get('purchase_price', 0))).border = border
            ws.cell(row=row, column=8, value=float(asset.get('current_value', 0))).border = border
            ws.cell(row=row, column=9, value=asset.get('department_code', '')).border = border
            ws.cell(row=row, column=10, value=asset.get('responsible_person', '')).border = border
        
        # Настройка ширины колонок
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Сохраняем файл
        filename = f"asset_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)
        wb.save(filepath)
        return filepath
    
    @staticmethod
    def _export_asset_pdf(data: Dict[str, Any]) -> str:
        """Экспорт отчета по активам в PDF"""
        filename = f"asset_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(REPORTS_DIR, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        
        # Заголовок
        story.append(Paragraph(data.get('title', 'Отчет по активам'), title_style))
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(f"Дата: {data.get('generated_at', datetime.now().isoformat())}", styles['Normal']))
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(f"Всего активов: {data.get('total_assets', 0)}", styles['Normal']))
        story.append(Spacer(1, 1*cm))
        
        # Таблица
        table_data = [
            ['№', 'Инв. номер', 'Наименование', 'Статус', 'Стоимость', 'Подразделение']
        ]
        
        for idx, asset in enumerate(data.get('assets', [])[:50], 1):
            table_data.append([
                str(idx),
                asset.get('inventory_number', '')[:20],
                asset.get('name', '')[:30],
                asset.get('status', ''),
                f"{float(asset.get('current_value', 0)):.2f}",
                asset.get('department_code', '')[:15],
            ])
        
        # Настройка таблицы
        table = Table(table_data, colWidths=[0.8*cm, 3*cm, 5*cm, 2*cm, 2.5*cm, 2.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(table)
        doc.build(story)
        return filepath
    
    @staticmethod
    def _export_asset_json(data: Dict[str, Any]) -> str:
        """Экспорт отчета по активам в JSON"""
        filename = f"asset_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(REPORTS_DIR, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        return filepath
    
    @staticmethod
    def _export_depreciation_excel(data: Dict[str, Any]) -> str:
        """Экспорт отчета по амортизации в Excel"""
        wb = Workbook()
        ws = ReportExporter._get_worksheet(wb)
        ws.title = "Амортизация"
        
        # Стили
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = data.get('title', 'Отчет по амортизации')
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:I2')
        ws['A2'] = f"Дата генерации: {data.get('generated_at', datetime.now().isoformat())}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A3:I3')
        ws['A3'] = f"Всего записей: {data.get('total_records', 0)}"
        ws['A3'].alignment = Alignment(horizontal="center")
        
        # Заголовки таблицы
        headers = ['№', 'ID актива', 'Период', 'Сумма', 'Накоплено', 'Стоимость до', 'Стоимость после', 'Ставка', 'Метод']
        row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Данные
        for idx, record in enumerate(data.get('records', []), 1):
            row += 1
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=record.get('asset_id', '')).border = border
            ws.cell(row=row, column=3, value=record.get('period', '')).border = border
            ws.cell(row=row, column=4, value=float(record.get('amount', 0))).border = border
            ws.cell(row=row, column=5, value=float(record.get('accumulated', 0))).border = border
            ws.cell(row=row, column=6, value=float(record.get('book_value_before', 0))).border = border
            ws.cell(row=row, column=7, value=float(record.get('book_value_after', 0))).border = border
            ws.cell(row=row, column=8, value=float(record.get('rate', 0))).border = border
            ws.cell(row=row, column=9, value=record.get('method', '')).border = border
        
        # Настройка ширины колонок
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 16
        
        filename = f"depreciation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)
        wb.save(filepath)
        return filepath
    
    @staticmethod
    def _export_depreciation_pdf(data: Dict[str, Any]) -> str:
        """Экспорт отчета по амортизации в PDF"""
        filename = f"depreciation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(REPORTS_DIR, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        
        story.append(Paragraph(data.get('title', 'Отчет по амортизации'), title_style))
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(f"Дата: {data.get('generated_at', datetime.now().isoformat())}", styles['Normal']))
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(f"Всего записей: {data.get('total_records', 0)}", styles['Normal']))
        story.append(Spacer(1, 1*cm))
        
        # Таблица
        table_data = [
            ['№', 'Актив', 'Период', 'Сумма', 'Ставка']
        ]
        
        for idx, record in enumerate(data.get('records', [])[:50], 1):
            table_data.append([
                str(idx),
                str(record.get('asset_id', '')),
                record.get('period', '')[:20],
                f"{float(record.get('amount', 0)):.2f}",
                f"{float(record.get('rate', 0))}%",
            ])
        
        table = Table(table_data, colWidths=[0.8*cm, 2*cm, 4*cm, 2.5*cm, 1.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(table)
        doc.build(story)
        return filepath
    
    @staticmethod
    def _export_depreciation_json(data: Dict[str, Any]) -> str:
        """Экспорт отчета по амортизации в JSON"""
        filename = f"depreciation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(REPORTS_DIR, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        return filepath
    
    @staticmethod
    def _export_inventory_excel(data: Dict[str, Any]) -> str:
        """Экспорт инвентаризационного отчета в Excel"""
        wb = Workbook()
        ws = ReportExporter._get_worksheet(wb)
        ws.title = "Инвентаризация"
        
        # Заголовок
        ws.merge_cells('A1:B1')
        ws['A1'] = data.get('title', 'Инвентаризационный отчет')
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws['A2'] = f"Дата: {data.get('generated_at', datetime.now().isoformat())}"
        
        ws['A4'] = "Статистика"
        ws['A4'].font = Font(bold=True, size=14)
        
        stats = data.get('stats', {})
        row = 6
        for key, value in stats.items():
            ws.cell(row=row, column=1, value=key.capitalize()).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)
        wb.save(filepath)
        return filepath
    
    @staticmethod
    def _export_inventory_pdf(data: Dict[str, Any]) -> str:
        """Экспорт инвентаризационного отчета в PDF"""
        filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(REPORTS_DIR, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        
        story.append(Paragraph(data.get('title', 'Инвентаризационный отчет'), title_style))
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(f"Дата: {data.get('generated_at', datetime.now().isoformat())}", styles['Normal']))
        story.append(Spacer(1, 1*cm))
        
        story.append(Paragraph("Статистика", styles['Heading2']))
        story.append(Spacer(1, 0.3*cm))
        
        stats = data.get('stats', {})
        for key, value in stats.items():
            story.append(Paragraph(f"{key.capitalize()}: {value}", styles['Normal']))
            story.append(Spacer(1, 0.2*cm))
        
        doc.build(story)
        return filepath
    
    @staticmethod
    def _export_inventory_json(data: Dict[str, Any]) -> str:
        """Экспорт инвентаризационного отчета в JSON"""
        filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(REPORTS_DIR, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        return filepath